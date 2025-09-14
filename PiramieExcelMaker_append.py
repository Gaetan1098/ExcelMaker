"""
PiramieExcelMaker_append.py
---------------------------------
Append a monthly purchases report into the MASTER's 'APR Bundle' sheet
*without* modifying headers or existing rows. Copies column-by-column
according to the mapping below, creates a backup first, expands the
sheet's AutoFilter / Table range, and normalizes 'Purchase Date' to real
datetimes so Excel sorting works.

Public API (kept the same for GUI compatibility):
    ingest_month_into_apr_bundle(master_xlsm_path, month_report_path) -> dict
Returns (GUI-friendly):
    {
        "master_backup": <path>,
        "updated_master": <path>,
        "rows_before": int,
        "rows_added": int,
        "rows_after": int,
        "sheet": "APR Bundle",
        # extra debug keys (not used by GUI, but handy):
        "unmapped_monthly_columns": [...],
        "unresolved_targets": {monthly_col: [APR target names...]},
        "resolved_map": {monthly_col: "ExcelColumnLetter"}
    }

Mapping (Monthly -> APR Bundle):
  Cust Name       -> CUSTOMER_NAME
  Cust Type       -> CUSTOMER_TYPE
  MSISDN          -> MSISDN
  Purchase Date   -> Purchase Date
  Prod Name       -> PRODUCT_NAME
  Amount          -> PURCHASE_AMT
  Package Status  -> STAT
  API Credit Type -> (API Credit Type | API  Credit Type)  # both spellings
  Prod Code       -> PRODUCT_ID
  CRTR_ID         -> CONTRACT_ID
"""

from __future__ import annotations

import os
import re
import sys
import shutil
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

APR_SHEET = "APR Bundle"
DEDUPE_FIELDS = ["MSISDN", "Purchase Date", "PRODUCT_NAME", "PURCHASE_AMT", "CONTRACT_ID", "PRODUCT_ID"]

# Monthly -> APR target mapping (APR targets are matched case/space-insensitively)
COLUMN_MAP: Dict[str, List[str]] = {
    "Cust Name": ["CUSTOMER_NAME"],
    "Cust Type": ["CUSTOMER_TYPE"],
    "MSISDN": ["MSISDN"],
    "Purchase Date": ["Purchase Date"],
    "Prod Name": ["PRODUCT_NAME"],
    "Amount": ["PURCHASE_AMT"],
    "Package Status": ["STAT"],
    "API Credit Type": ["API Credit Type", "API  Credit Type"],  # 1 or 2 spaces
    "Prod Code": ["PRODUCT_ID"],
    "CRTR_ID": ["CONTRACT_ID"],
}

# If the monthly report has a fixed sheet name, set it here; else first sheet
MONTH_SHEET_NAME: Optional[str] = None
# Your monthly report's headers are on *row 4* (0-indexed header=3)
MONTH_HEADER_ROW = 3


# -------------------- utilities --------------------

def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _backup_master(master_path: str) -> str:
    backup_dir = os.path.join(os.path.dirname(master_path), "_backups")
    os.makedirs(backup_dir, exist_ok=True)
    name, ext = os.path.splitext(os.path.basename(master_path))
    backup_path = os.path.join(backup_dir, f"{name}_backup_{_timestamp()}{ext}")
    shutil.copy2(master_path, backup_path)
    return backup_path


def _norm(s: str) -> str:
    """normalize text for header comparisons: collapse spaces, lowercase, strip"""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _read_month_file(month_path: str) -> pd.DataFrame:
    """
    Read the month purchases report (header is on row 4), then normalize/derive.
    """
    try:
        # header=3 => Excel row 4 contains the real headers
        raw = pd.read_excel(month_path, sheet_name=0, header=3)
    except Exception as e:
        raise RuntimeError(f"Failed to read monthly file '{month_path}': {e}")

    raw = _normalize_columns(raw)
    raw = _coerce_and_derive(raw)

    return raw



def _build_apr_header_index(ws) -> Dict[str, int]:
    """Row 1 headers in APR sheet -> col index (1-based), normalized."""
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header_map[_norm(cell.value)] = col_idx
    return header_map


def _resolve_target_col(apr_header_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
    """Choose the first matching APR target name (case/space-insensitive)."""
    for name in candidates:
        key = _norm(name)
        if key in apr_header_map:
            return apr_header_map[key]
    return None


def _coerce_for_excel(val):
    """Convert pandas NA/NaT to None; strip strings; leave datetimes as datetime."""
    if pd.isna(val):
        return None
    if isinstance(val, str):
        return val.strip()
    return val


def _last_used_row(ws, from_row=2, to_col=None):
    """Find last row that has any value in columns 1..to_col (defaults to header width)."""
    if to_col is None:
        to_col = max((c.column for c in ws[1] if c.value is not None), default=ws.max_column)
    last = 1
    for r in range(from_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, to_col + 1)):
            last = r
    return last, to_col


def _expand_filters_and_tables(ws):
    """
    Expand AutoFilter ref and any Excel Table(s) to include all appended rows.
    Assumes headers are on row 1 and data begins on row 2.
    """
    last_row, last_col = _last_used_row(ws, from_row=2)
    end_col_letter = get_column_letter(last_col)

    # 1) AutoFilter
    try:
        if ws.auto_filter and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A1:{end_col_letter}{last_row}"
    except Exception:
        pass

    # 2) Structured Tables
    tbls = []
    try:
        tbls = list(getattr(ws, "tables", {}).values())  # new openpyxl
    except Exception:
        pass
    if not tbls:
        tbls = getattr(ws, "_tables", []) or []  # old openpyxl

    for tbl in tbls:
        try:
            start, _ = tbl.ref.split(":")
            start_row = int("".join(filter(str.isdigit, start)))
            if start_row == 1:
                # keep existing width; only extend downward
                _, end = tbl.ref.split(":")
                end_col_letter_current = "".join(filter(str.isalpha, end))
                tbl.ref = f"{''.join(filter(str.isalpha, start))}1:{end_col_letter_current}{last_row}"
        except Exception:
            continue


def _find_header_col(ws, header_text: str) -> Optional[int]:
    target = _norm(header_text)
    for c, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        if _norm(cell.value) == target:
            return c
    return None


def _normalize_purchase_date_column(ws):
    """Coerce ALL 'Purchase Date' cells to true datetimes; set uniform number format."""
    col = _find_header_col(ws, "Purchase Date")
    if not col:
        return
    last_row, _ = _last_used_row(ws, from_row=2)
    for r in range(2, last_row + 1):
        cell = ws.cell(row=r, column=col)
        v = cell.value
        if v is None or isinstance(v, datetime):
            # already empty or a real datetime
            continue
        try:
            dt = pd.to_datetime(str(v), errors="raise")
            if pd.notna(dt):
                cell.value = dt.to_pydatetime()
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        except Exception:
            # leave as-is if not parseable
            pass

# Treat these columns as the row identity (use only those that exist in APR headers / month df)
DEDUPE_FIELDS = ["MSISDN", "Purchase Date", "PRODUCT_NAME", "PURCHASE_AMT", "CONTRACT_ID", "PRODUCT_ID"]

def _read_existing_keys(ws) -> set[tuple]:
    """
    Build a set of identity keys from existing APR rows using DEDUPE_FIELDS.
    Keys are normalized (dates to second, amounts to float, strings stripped).
    """
    # which of the identity fields actually exist on the APR sheet?
    present = []
    for f in DEDUPE_FIELDS:
        col = _find_header_col(ws, f)
        if col:
            present.append((f, col))
    if not present:
        return set()

    keys = set()
    last_row, _ = _last_used_row(ws, from_row=2)
    for r in range(2, last_row + 1):
        parts = []
        for f, col in present:
            v = ws.cell(row=r, column=col).value
            if v is None:
                parts.append(None)
                continue
            if f == "Purchase Date":
                dt = pd.to_datetime(v, errors="coerce")
                parts.append(None if pd.isna(dt) else dt.to_pydatetime().replace(microsecond=0))
            elif f == "PURCHASE_AMT":
                try:
                    parts.append(float(str(v).replace(",", "").replace("$", "").strip()))
                except Exception:
                    parts.append(v)
            else:
                parts.append(str(v).strip() if isinstance(v, str) else v)
        keys.add(tuple(parts))
    return keys

def _make_row_key(series: pd.Series) -> tuple:
    """
    Build a normalized key from a monthly DataFrame row using DEDUPE_FIELDS.
    """
    parts = []
    for f in DEDUPE_FIELDS:
        if f not in series.index or pd.isna(series[f]):
            parts.append(None)
            continue
        v = series[f]
        if f == "Purchase Date":
            dt = pd.to_datetime(v, errors="coerce")
            parts.append(None if pd.isna(dt) else dt.to_pydatetime().replace(microsecond=0))
        elif f == "PURCHASE_AMT":
            try:
                parts.append(float(str(v).replace(",", "").replace("$", "").strip()))
            except Exception:
                parts.append(v)
        else:
            parts.append(str(v).strip() if isinstance(v, str) else v)
    return tuple(parts)

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize monthly headers from row 4 to a canonical set so later mapping works.
    We keep the 'Monthly' label names from COLUMN_MAP keys on the DataFrame
    (e.g., 'Prod Name', 'Amount', etc.).
    """
    df = df.copy()

    # Build a normalization of current columns -> canonical monthly labels
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s)).strip().lower()

    # aliases (normalized) -> canonical monthly label
    aliases = {
        "cust name": "Cust Name",
        "customer name": "Cust Name",

        "cust type": "Cust Type",
        "customer type": "Cust Type",

        "msisdn": "MSISDN",

        "purchase date": "Purchase Date",
        "date": "Purchase Date",

        "prod name": "Prod Name",
        "product name": "Prod Name",

        "amount": "Amount",
        "purchase amount": "Amount",
        "purchase amt": "Amount",

        "package status": "Package Status",
        "stat": "Package Status",

        # collapse multiple spaces automatically; both map to same
        "api credit type": "API Credit Type",

        "prod code": "Prod Code",
        "product id": "Prod Code",

        "crtr_id": "CRTR_ID",
        "crtr id": "CRTR_ID",
        "contract id": "CRTR_ID",
    }

    new_cols = {}
    for c in df.columns:
        nc = norm(c)
        if nc in aliases:
            new_cols[c] = aliases[nc]
        else:
            # keep as-is for pass-through columns we don't explicitly map
            new_cols[c] = c
    df = df.rename(columns=new_cols)

    # Drop completely empty columns (often 'Unnamed: ...' from Excel)
    empty_cols = [c for c in df.columns if df[c].isna().all()]
    if empty_cols:
        df = df.drop(columns=empty_cols)

    return df


def _coerce_and_derive(df: pd.DataFrame) -> pd.DataFrame:
    """
    Make sure the monthly DF has the APR-style columns used by the dedupe key:
      PRODUCT_NAME, PURCHASE_AMT, PRODUCT_ID, CONTRACT_ID, MSISDN, Purchase Date.
    Also coerce types (dates & numeric amounts) for reliable comparison.
    """
    df = df.copy()

    # Coerce Purchase Date to datetime (keep original monthly label for mapping)
    if "Purchase Date" in df.columns:
        df["Purchase Date"] = pd.to_datetime(df["Purchase Date"], errors="coerce")

    # MSISDN as string stripped (safe for leading zeros etc.)
    if "MSISDN" in df.columns:
        df["MSISDN"] = df["MSISDN"].astype(str).str.strip()
        # Replace 'nan' artifacts from astype(str)
        df.loc[df["MSISDN"].str.lower().isin(["nan", "none", ""]), "MSISDN"] = pd.NA

    # Amount -> numeric PURCHASE_AMT (APR-style) + keep "Amount" for mapping to APR column
    if "Amount" in df.columns:
        amt = df["Amount"].astype(str).str.replace(",", "", regex=False).str.replace("$", "", regex=False).str.strip()
        df["PURCHASE_AMT"] = pd.to_numeric(amt, errors="coerce")
    else:
        df["PURCHASE_AMT"] = pd.NA

    # Map monthly columns into APR-style names so dedupe has what it needs
    df["PRODUCT_NAME"] = df["Prod Name"] if "Prod Name" in df.columns else pd.NA
    df["PRODUCT_ID"] = df["Prod Code"] if "Prod Code" in df.columns else pd.NA
    df["CONTRACT_ID"] = df["CRTR_ID"] if "CRTR_ID" in df.columns else pd.NA

    # Ensure the dedupe fields exist even if missing in the source (filled with NA)
    for col in ["MSISDN", "Purchase Date", "PRODUCT_NAME", "PURCHASE_AMT", "CONTRACT_ID", "PRODUCT_ID"]:
        if col not in df.columns:
            df[col] = pd.NA

    return df

# -------------------- main API (kept name/signature) --------------------

def ingest_month_into_apr_bundle(master_xlsm_path: str, month_report_path: str) -> dict:
    """
    Append monthly columns into APR Bundle (column-by-column, no header rewrite).
    Creates a timestamped backup before writing.
    Returns GUI-friendly summary: rows_before / rows_added / rows_after.
    """
    if not os.path.isfile(master_xlsm_path):
        raise FileNotFoundError(f"Master not found: {master_xlsm_path}")
    if not os.path.isfile(month_report_path):
        raise FileNotFoundError(f"Month file not found: {month_report_path}")

    # Load month DF with tolerant headers and header row 4
    month_df = _read_month_file(month_report_path)

    wb = load_workbook(master_xlsm_path, keep_vba=True)
    if APR_SHEET not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"'{APR_SHEET}' sheet not found in master.")
    ws = wb[APR_SHEET]

    # Build a set of existing identity keys to prevent duplicate appends
    existing_keys = _read_existing_keys(ws)

    # Rows before
    last_used_before, header_width = _last_used_row(ws, from_row=2)
    rows_before = max(0, last_used_before - 1)  # data starts at row 2

    # Build APR header map
    apr_header_map = _build_apr_header_index(ws)

    # Resolve mapping Monthly -> APR col index
    resolved_map: Dict[str, int] = {}
    skipped_targets: Dict[str, List[str]] = {}
    missing_monthly: List[str] = []
    for monthly_label, targets in COLUMN_MAP.items():
        if monthly_label not in month_df.columns:
            missing_monthly.append(monthly_label)
            continue
        col_idx = _resolve_target_col(apr_header_map, targets)
        if col_idx is None:
            skipped_targets[monthly_label] = targets
        else:
            resolved_map[monthly_label] = col_idx

    # If nothing resolved, still make a safe backup and return diagnostics
    backup_path = _backup_master(master_xlsm_path)

    if not resolved_map:
        wb.close()
        return {
            "master_backup": backup_path,
            "updated_master": master_xlsm_path,
            "rows_before": rows_before,
            "rows_added": 0,
            "rows_after": rows_before,
            "sheet": APR_SHEET,
            "unmapped_monthly_columns": missing_monthly,
            "unresolved_targets": skipped_targets,
            "resolved_map": {},
        }

    # First empty row after current data
    start_row = last_used_before + 1

    # Append values cell-by-cell with de-duplication
    n_rows = len(month_df)
    written = 0
    duplicates_skipped = 0

    for i in range(n_rows):
        # Build key for the current monthly row (uses any matching DEDUPE_FIELDS present)
        row_key = _make_row_key(month_df.iloc[i])
        if row_key in existing_keys:
            duplicates_skipped += 1
            continue  # skip duplicates already in APR

        target_row = start_row + written  # compact placement (no gaps)
        row_has_any = False
        for m_col, col_idx in resolved_map.items():
            val = _coerce_for_excel(month_df.iloc[i][m_col])

            # Amount -> numeric if possible
            if m_col == "Amount" and isinstance(val, str):
                val = re.sub(r"[,$]", "", val).strip()
                try:
                    val = float(val)
                except Exception:
                    pass

            # Purchase Date -> real datetime
            if m_col == "Purchase Date":
                val = pd.to_datetime(val, errors="coerce")
                val = None if pd.isna(val) else val.to_pydatetime().replace(microsecond=0)

            if val is not None:
                row_has_any = True
            ws.cell(row=target_row, column=col_idx, value=val)

        if row_has_any:
            written += 1
            existing_keys.add(row_key)  # also prevents in-file duplicates

    # Expand table & autofilter; normalize dates column for correct sort
    _expand_filters_and_tables(ws)
    _normalize_purchase_date_column(ws)

    wb.save(master_xlsm_path)
    wb.close()

    # Rows after
    last_used_after, _ = _last_used_row(ws, from_row=2)
    rows_after = max(0, last_used_after - 1)
    rows_added = written

    return {
        "master_backup": backup_path,
        "updated_master": master_xlsm_path,
        "rows_before": rows_before,
        "rows_added": rows_added,
        "rows_after": rows_after,
        "sheet": APR_SHEET,
        "dedupe_skipped": duplicates_skipped,  # <-- new
        "unmapped_monthly_columns": [c for c in COLUMN_MAP.keys() if c not in month_df.columns],
        "unresolved_targets": skipped_targets,
        "resolved_map": {k: get_column_letter(v) for k, v in resolved_map.items()}, 
        }


# ------------- CLI -------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Usage:\n"
            "  python PiramieExcelMaker_append.py <MASTER_XLSM_PATH> <MONTH_REPORT_PATH>\n"
        )
        sys.exit(1)

    master = sys.argv[1]
    month = sys.argv[2]
    summary = ingest_month_into_apr_bundle(master, month)
    print("Ingest complete:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
