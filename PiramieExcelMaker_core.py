def process_excel_file(excel_path):
    import sys
    import os
    import tkinter as tk
    from tkinter import filedialog
    import pandas as pd
    import openpyxl 
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    import shutil
    from datetime import datetime

    #use MoloMolo MSISDN to identify data from raw data
    #use APR Bundle with MSISDN to calculate data to add to 3rd sheet

    #load excel file
    # `excel_path` will be injected from GUI wrapper
    # Ensure it's declared at the top if needed:
    # excel_path = ""
    APR_dataframe = pd.read_excel(excel_path, sheet_name='APR Bundle')
    Molo_dataframe = pd.read_excel(excel_path, sheet_name='Molo Molo')
    print(Molo_dataframe[["Name"]].head(10))

    #Backup 
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.dirname(excel_path)
    backup_path = os.path.join(backup_dir, f"Comission_calculations_Backup_{timestamp}.xlsm")
    shutil.copyfile(excel_path, backup_path)
    print(f"backup created: {backup_path}")

    #Prepare MSISDNs
    Molo_MSISDNs = Molo_dataframe[["MSISDN", "Reg Date"]].dropna(subset=['MSISDN']).drop_duplicates()
    valid_MSISDNs = Molo_MSISDNs["MSISDN"].unique()

    #Count purchases and total ammount of topup
    grouped = APR_dataframe.groupby("MSISDN")["PURCHASE_AMT"].agg(["count", "sum"]).reset_index()
    grouped.columns = ["MSISDN", "TOPUP COUNT", "TOPUP AMOUNT"]

    #Sum Ecesia-credited ammounts
    encisia_dataframe = APR_dataframe[APR_dataframe["API  Credit Type"] == "encisia"]
    encisia_sum = encisia_dataframe.groupby("MSISDN")["PURCHASE_AMT"].sum().reset_index()
    encisia_sum.columns = ["MSISDN", "ENCISIA"]

    #Compute HQ
    combined = pd.merge(grouped, encisia_sum, on="MSISDN", how="left")
    combined["ENCISIA"] = combined["ENCISIA"].fillna(0)
    combined["HQ"] = combined["TOPUP AMOUNT"] - combined["ENCISIA"]

    #Names
    names = APR_dataframe[["MSISDN", "CUSTOMER_NAME"]].drop_duplicates()
    combined = pd.merge(combined, names, on="MSISDN", how="left")

    #Data Type
    MSISDN_to_produt = APR_dataframe.drop_duplicates(subset=["MSISDN"])[["MSISDN", "PRODUCT_NAME"]]

    #Merge with Molo sheet for reg dates
    final = pd.merge(Molo_MSISDNs, combined, on='MSISDN', how='left')

    #Merge with data type
    final = pd.merge(final, MSISDN_to_produt, on="MSISDN", how="left")
    final = final.rename(columns={"PRODUCT_NAME": "DATA Type"})

    #Monthly Summary
    APR_filtered = APR_dataframe[APR_dataframe["MSISDN"].isin(valid_MSISDNs)].copy()
    APR_filtered["Month"] = pd.to_datetime(APR_filtered["Purchase Date"]).dt.to_period('M')
    available_months = sorted(APR_filtered["Month"].unique())

    monthly_summary = (
        APR_filtered.groupby(["MSISDN", "Month"])["PURCHASE_AMT"]
        .agg(["sum", "count"])
        .reset_index()
    )

    monthly_encisia = APR_filtered[APR_filtered["API  Credit Type"] == "encisia"]
    monthly_encisia = monthly_encisia.groupby(["MSISDN", "Month"])["PURCHASE_AMT"].sum().reset_index()
    monthly_encisia.columns = ["MSISDN", "Month", "ENCISIA"]

    monthly_summary = pd.merge(monthly_summary, monthly_encisia, on=["MSISDN", "Month"], how="left")
    monthly_summary["ENCISIA"] = monthly_summary["ENCISIA"].fillna(0)
    monthly_summary["HQ"] = monthly_summary["sum"] - monthly_summary["ENCISIA"]

    metrics_order = ['sum', 'count', 'ENCISIA', 'HQ']
    monthly_pivot = monthly_summary.pivot(index="MSISDN", 
                                        columns="Month", 
                                        values=metrics_order)

    monthly_pivot.columns = [f"{month.strftime('%B %Y')} {metric.upper()}" 
                            for metric in metrics_order 
                            for month in available_months]
    monthly_pivot = monthly_pivot.reset_index()


    final = pd.merge(final, monthly_pivot, on="MSISDN", how="left")
    final.fillna(0, inplace=True)


    #Reorder and add Data Type
    static_columns = ["MSISDN", "CUSTOMER_NAME", "Reg Date", "TOPUP COUNT", "TOPUP AMOUNT", "ENCISIA", "HQ", "DATA Type"]
    monthly_columns = [col for col in final.columns if col not in static_columns]

    def sort_monthly_columns(col):
        parts = col.split()
        if len(parts) >= 3:
            month_str = ' '.join(parts[:2])
            metric = ' '.join(parts[2:]).lower()  # Convert metric to lowercase
            try:
                return (pd.to_datetime(month_str), metrics_order.index(metric))
            except ValueError:
                # If metric not found, place it at the end
                return (pd.to_datetime(month_str), len(metrics_order))
        return (pd.to_datetime('1900-01'), 0)

    monthly_columns_sorted = sorted(monthly_columns, key=sort_monthly_columns)

    final = final[static_columns + monthly_columns_sorted]


    #Writing Molo Molo Auto
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)
    if "Molo Molo Auto" in wb.sheetnames:
        del wb["Molo Molo Auto"]
    ws = wb.create_sheet("Molo Molo Auto")

    ws.freeze_panes = 'A3'

    #preparing headers
    main_headers = []
    sub_headers = []
    for col in final.columns:
        if any(month.strftime("%B %Y") in col for month in available_months):
            parts = col.split(" ")
            main = " ".join(parts[:2])
            sub = " ".join(parts[2:])
            main_headers.append(main)
            sub_headers.append(sub)
        else:
            main_headers.append(col)
            sub_headers.append("")


    # Writing headers
    col_index = 1
    current_month = None
    month_start_col = 1

    for col in final.columns:
        if any(month.strftime("%B %Y") in col for month in available_months):
            parts = col.split()
            month_name = ' '.join(parts[:2])
            metric = ' '.join(parts[2:])
            
            if month_name != current_month:
                # New month - finish previous month's header if exists
                if current_month is not None:
                    ws.merge_cells(start_row=1, start_column=month_start_col, 
                                end_row=1, end_column=col_index-1)
                    ws.cell(row=1, column=month_start_col, 
                        value=current_month).alignment = Alignment(horizontal="center")
                
                # Start new month
                current_month = month_name
                month_start_col = col_index
            
            # Write subheader
            ws.cell(row=2, column=col_index, value=metric)
            col_index += 1
        else:
            # Non-monthly column
            if current_month is not None:
                # Finish the current month header
                ws.merge_cells(start_row=1, start_column=month_start_col, 
                            end_row=1, end_column=col_index-1)
                ws.cell(row=1, column=month_start_col, 
                    value=current_month).alignment = Alignment(horizontal="center")
                current_month = None
            
            # Write regular header
            ws.merge_cells(start_row=1, start_column=col_index, 
                        end_row=2, end_column=col_index)
            ws.cell(row=1, column=col_index, value=col)
            col_index += 1

    # Finish any remaining month header
    if current_month is not None:
        ws.merge_cells(start_row=1, start_column=month_start_col, 
                    end_row=1, end_column=col_index-1)
        ws.cell(row=1, column=month_start_col, 
            value=current_month).alignment = Alignment(horizontal="center")

    # Writing data 
    for row_index, row in final.iterrows():
        for col_index, val in enumerate(row, 1):
            ws.cell(row=row_index + 3, column=col_index, value=val)

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    #save file
    wb.save(excel_path)

    print(f"Molo Molo Auto sheet generated: {excel_path}")

if __name__ == "__main__":
    # test/debug mode
    process_excel_file("Commission_Calculations_April_2025_Gaetan.xlsm")